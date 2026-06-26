import random
from pathlib import Path
from typing import Iterable

FILE_NAME = "LO.xlsx"
FILE_PATH = "Files/Mancia"


def _iter_formatted_paragraphs(sheet) -> Iterable[dict[str, str]]:
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        text_value = row[0] if len(row) > 0 else None
        title_value = row[1] if len(row) > 1 else None
        page_value = row[3] if len(row) > 3 else None
        text = str(text_value).strip() if text_value is not None else ""
        title = str(title_value).strip() if title_value is not None else ""
        page = str(page_value).strip() if page_value is not None else ""
        if text and title:
            yield {"paragraph": text, "page": page}


def get_random_paragraph(term: str = "") -> dict:
    _ = term  # compatibility with existing call sites/signature
    file_path = Path(__file__).resolve().parents[1] / FILE_PATH / FILE_NAME
    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    try:
        import openpyxl  # type: ignore
    except Exception as exc:
        raise RuntimeError("Dependency 'openpyxl' is not available in backend.") from exc

    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        paragraphs = list(_iter_formatted_paragraphs(sheet))
    finally:
        workbook.close()

    if not paragraphs:
        raise ValueError(f"No valid paragraphs found in file: {FILE_NAME}")

    total_paragraphs = len(paragraphs)
    random_index = random.randint(0, total_paragraphs - 1)
    selected_paragraph = paragraphs[random_index]

    return {
        "paragraph": selected_paragraph["paragraph"],
        "page": selected_paragraph["page"],
        "paragraph_number": random_index + 1,
        "total_paragraphs": total_paragraphs,
        "source": file_path.name,
    }
