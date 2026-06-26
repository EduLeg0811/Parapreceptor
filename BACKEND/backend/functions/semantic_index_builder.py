from __future__ import annotations

import json
import os
import tempfile
import time
from pathlib import Path
from typing import Any

import numpy as np
import openpyxl  # type: ignore
import requests

try:
    from backend.functions.semantic_chunking import (
        DEFAULT_CHUNK_MAX_CHARS,
        DEFAULT_CHUNK_MIN_CHARS,
        DEFAULT_CHUNK_TARGET_CHARS,
        rechunk_semantic_rows,
    )
    from backend.functions.semantic_index_calibration import build_calibration_payload, compute_similarity_stats, recommend_min_score
    from backend.functions.source_records import load_source_document, source_path_for, strip_markdown
except Exception:
    from functions.semantic_chunking import (
        DEFAULT_CHUNK_MAX_CHARS,
        DEFAULT_CHUNK_MIN_CHARS,
        DEFAULT_CHUNK_TARGET_CHARS,
        rechunk_semantic_rows,
    )
    from functions.semantic_index_calibration import build_calibration_payload, compute_similarity_stats, recommend_min_score
    from functions.source_records import load_source_document, source_path_for, strip_markdown


EMBEDDINGS_API_URL = "https://api.openai.com/v1/embeddings"
EMBED_BATCH_SIZE = 64
_SESSION = requests.Session()
ROOT_DIR = Path(__file__).resolve().parents[2]


def _utc_iso_now() -> str:
    return time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())


def _load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _sanitize_cell(value: Any) -> str:
    return str(value or "").replace("\u00A0", " ").replace("\r", " ").replace("\n", " ").strip()


def _resolve_source_path(raw_path: str) -> Path:
    source_path = Path(raw_path)
    if source_path.is_absolute():
        return source_path
    return (ROOT_DIR / source_path).resolve()


def _write_json_atomic(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_path = tempfile.mkstemp(dir=str(path.parent), prefix=f"{path.name}.", suffix=".tmp")
    os.close(fd)
    try:
        Path(tmp_path).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        Path(tmp_path).replace(path)
    finally:
        tmp_file = Path(tmp_path)
        if tmp_file.exists():
            tmp_file.unlink(missing_ok=True)


def _write_npy_atomic(path: Path, array: np.ndarray) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_path = tempfile.mkstemp(dir=str(path.parent), prefix=f"{path.name}.", suffix=".tmp")
    os.close(fd)
    try:
        with open(tmp_path, "wb") as handle:
            np.save(handle, array)
        Path(tmp_path).replace(path)
    finally:
        tmp_file = Path(tmp_path)
        if tmp_file.exists():
            tmp_file.unlink(missing_ok=True)


def _normalize_rows(vectors: np.ndarray) -> np.ndarray:
    if vectors.ndim != 2:
        raise ValueError("Embeddings invalidos para indexacao semantica.")
    norms = np.linalg.norm(vectors, axis=1, keepdims=True)
    norms = np.where(norms > 0, norms, 1.0)
    return vectors / norms


def _embed_batch(texts: list[str], api_key: str, model: str) -> np.ndarray:
    response = _SESSION.post(
        EMBEDDINGS_API_URL,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        json={
            "model": model,
            "input": texts,
        },
        timeout=120,
    )
    if not response.ok:
        raise RuntimeError(f"Falha ao gerar embeddings: HTTP {response.status_code} {response.text}")
    payload = response.json()
    data = payload.get("data") or []
    if len(data) != len(texts):
        raise RuntimeError("Quantidade de embeddings retornada difere da quantidade de textos.")
    return np.asarray([item.get("embedding") or [] for item in data], dtype=np.float32)


def _embed_rows(rows: list[dict[str, Any]], api_key: str, model: str, batch_size: int = EMBED_BATCH_SIZE) -> np.ndarray:
    texts = [str(row.get("embedding_text") or strip_markdown(str(row.get("text") or ""))).strip() for row in rows]
    if not texts:
        return np.zeros((0, 0), dtype=np.float32)

    batches: list[np.ndarray] = []
    for start in range(0, len(texts), max(1, int(batch_size or EMBED_BATCH_SIZE))):
        batch = texts[start:start + max(1, int(batch_size or EMBED_BATCH_SIZE))]
        batches.append(_embed_batch(batch, api_key=api_key, model=model))
    return _normalize_rows(np.vstack(batches))


def _load_rows_from_xlsx(source_path: Path, manifest: dict[str, Any]) -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    try:
        sheet_name = str(manifest.get("sheet_name") or "").strip()
        if sheet_name and sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook[workbook.sheetnames[0]]

        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = [str(col).strip() if col is not None else "" for col in header_row]
        normalized_headers = [header.lower() for header in headers]
        text_column = str(manifest.get("text_column") or "text").strip().lower()
        metadata_columns = [str(item).strip() for item in (manifest.get("metadata_columns") or []) if str(item).strip()]
        metadata_columns_norm = {column.lower(): column for column in metadata_columns}

        rows: list[dict[str, Any]] = []
        for row_index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not values:
                continue

            row_map: dict[str, str] = {}
            for idx, value in enumerate(values):
                key = normalized_headers[idx] if idx < len(normalized_headers) and normalized_headers[idx] else f"col_{idx + 1}"
                row_map[key] = _sanitize_cell(value)

            text = row_map.get(text_column, "").strip()
            if not text:
                continue

            metadata = {
                target_key: row_map.get(source_key, "").strip()
                for source_key, target_key in metadata_columns_norm.items()
                if row_map.get(source_key, "").strip()
            }
            rows.append({
                "row": row_index,
                "text": text,
                "metadata": metadata,
            })

        return rows
    finally:
        workbook.close()


def _load_rows_for_rebuild(
    index_dir: Path,
    manifest: dict[str, Any],
    *,
    require_source_file: bool,
) -> tuple[list[dict[str, Any]], str, str | None]:
    try:
        source_doc = load_source_document(index_dir.name)
        records = source_doc.get("records") if isinstance(source_doc, dict) else None
        if isinstance(records, list):
            return records, "source_json", None
    except FileNotFoundError:
        pass

    chunks_path = index_dir / "chunks.json"
    if chunks_path.exists():
        chunks = _load_json(chunks_path)
        if isinstance(chunks, list):
            return chunks, "chunks_json", None

    source_file = str(manifest.get("source_file") or "").strip()
    if source_file:
        source_path = _resolve_source_path(source_file)
        if source_path.exists():
            return _load_rows_from_xlsx(source_path, manifest), "source_file", None
        if require_source_file:
            raise FileNotFoundError(f"Arquivo-fonte do indice nao encontrado: {source_path}")

    if source_file:
        raise FileNotFoundError(f"Arquivo-fonte do indice nao encontrado: {_resolve_source_path(source_file)}")
    raise FileNotFoundError(f"Fonte JSON nao encontrada para rebuild do indice: {index_dir.name}")


def rebuild_semantic_index(
    index_dir: Path,
    *,
    api_key: str,
    model: str | None = None,
    output_dir: Path | None = None,
    batch_size: int = EMBED_BATCH_SIZE,
    target_chars: int = DEFAULT_CHUNK_TARGET_CHARS,
    max_chars: int = DEFAULT_CHUNK_MAX_CHARS,
    min_chars: int = DEFAULT_CHUNK_MIN_CHARS,
    require_source_file: bool = False,
) -> dict[str, Any]:
    manifest_path = index_dir / "manifest.json"
    if not manifest_path.exists():
        raise FileNotFoundError(f"Indice semantico incompleto: {index_dir}")

    manifest = _load_json(manifest_path)
    source_rows, rebuild_basis, warning = _load_rows_for_rebuild(
        index_dir,
        manifest,
        require_source_file=require_source_file,
    )

    resolved_model = (model or manifest.get("model") or "text-embedding-3-small").strip()
    index_label = str(manifest.get("index_label") or index_dir.name).strip()
    rebuilt_rows = rechunk_semantic_rows(
        source_rows,
        index_label=index_label,
        target_chars=target_chars,
        max_chars=max_chars,
        min_chars=min_chars,
    )
    if not rebuilt_rows:
        raise ValueError(f"Nenhum chunk gerado para o indice {index_dir.name}")

    embeddings = _embed_rows(rebuilt_rows, api_key=api_key, model=resolved_model, batch_size=batch_size)
    embeddings_to_disk = embeddings.astype(np.float16, copy=False)
    calibration_stats = compute_similarity_stats(embeddings)
    recommended_min_score = recommend_min_score(calibration_stats)

    target_dir = output_dir or index_dir
    target_dir.mkdir(parents=True, exist_ok=True)

    manifest["model"] = resolved_model
    manifest["embedding_dtype"] = "float16"
    manifest["dimensions"] = int(embeddings.shape[1])
    manifest["indexed_rows"] = len(rebuilt_rows)
    manifest["generated_at"] = _utc_iso_now()
    manifest["normalization"] = "l2"
    manifest["recommended_min_score"] = recommended_min_score
    manifest["score_calibration"] = build_calibration_payload(calibration_stats)
    manifest["rebuild_basis"] = rebuild_basis
    manifest["chunking"] = {
        "version": 1,
        "strategy": "structural_sentence_rechunk",
        "targetChars": int(target_chars),
        "maxChars": int(max_chars),
        "minChars": int(min_chars),
        "rebuiltFromRows": len(source_rows),
    }

    stored_rows = []
    for row in rebuilt_rows:
        stored_row = dict(row)
        stored_row.pop("embedding_text", None)
        stored_rows.append(stored_row)

    _write_npy_atomic(target_dir / "embeddings.npy", embeddings_to_disk)
    _write_json_atomic(target_dir / "manifest.json", manifest)
    _write_json_atomic(target_dir / "chunks.json", stored_rows)

    try:
        source_path = source_path_for(index_dir.name)
        source_doc = _load_json(source_path)
        if isinstance(source_doc, dict) and "semantic_rows" in source_doc:
            source_doc["semantic_rows"] = len(stored_rows)
            _write_json_atomic(source_path, source_doc)
    except FileNotFoundError:
        pass

    return {
        "index_id": index_dir.name,
        "rows_before": len(source_rows),
        "rows_after": len(stored_rows),
        "model": resolved_model,
        "recommended_min_score": recommended_min_score,
        "output_dir": str(target_dir),
        "rebuild_basis": rebuild_basis,
        "warning": warning,
    }
