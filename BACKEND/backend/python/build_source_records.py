from __future__ import annotations

import json
import sys
import time
from pathlib import Path
from typing import Any

import openpyxl  # type: ignore


PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from backend.functions.semantic_chunking import rechunk_semantic_rows  # noqa: E402


LEXICAL_DIR = PROJECT_ROOT / "backend" / "Files" / "Lexical"
SEMANTIC_DIR = PROJECT_ROOT / "backend" / "Files" / "Semantic"
SOURCE_DIR = PROJECT_ROOT / "backend" / "Files" / "Source"

FILE_STEM_TO_INDEX_ID = {
    "200TEAT": "200teat",
    "700EXP": "700exp",
    "CCG": "ccg",
    "DAC": "dac",
    "DUPLA": "dupla",
    "EC": "ec",
    "HSP": "hsp",
    "HSR": "hsr",
    "LO": "lo",
    "PROEXIS": "proexis",
    "PROJ": "proj",
    "QUEST": "quest",
    "TEMAS": "temas",
    "TNP": "tnp",
}

FILE_STEM_TO_BOOK_CODE = {
    "200TEAT": "TEAT",
    "700EXP": "EXP",
    "CCG": "CCG",
    "DAC": "DAC",
    "DUPLA": "MDE",
    "EC": "EC",
    "HSP": "HSP",
    "HSR": "HSR",
    "LO": "LO",
    "PROEXIS": "MP",
    "PROJ": "PROJ",
    "QUEST": "QUEST",
    "TEMAS": "TC",
    "TNP": "TNP",
}

EXTRA_JSON_SOURCES = [
    {
        "raw_path": SOURCE_DIR / "arlindo_mini_records.json",
        "index_id": "mini_arlindo",
        "index_label": "Minitertulia - Arlindo",
        "book_code": "MINI_ARLINDO",
        "file_stem": "MINI_ARLINDO",
        "output_name": "mini_arlindo.json",
        "metadata_columns": ["date", "cm", "quest", "pens", "number"],
    }
]


def _utc_iso_now() -> str:
    return time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())


def _load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _sanitize_cell(value: Any) -> str:
    return str(value or "").replace("\u00A0", " ").replace("\r", " ").replace("\n", " ").strip()


def _metadata_columns_for(index_id: str, headers: list[str]) -> list[str]:
    manifest_path = SEMANTIC_DIR / index_id / "manifest.json"
    if manifest_path.exists():
        manifest = _load_json(manifest_path)
        columns = [str(item).strip() for item in manifest.get("metadata_columns") or [] if str(item).strip()]
        if columns:
            return columns
    return [header for header in headers if header and header.lower() != "text"]


def _sheet_name_for(index_id: str, workbook: Any) -> str:
    manifest_path = SEMANTIC_DIR / index_id / "manifest.json"
    if manifest_path.exists():
        manifest = _load_json(manifest_path)
        sheet_name = str(manifest.get("sheet_name") or "").strip()
        if sheet_name and sheet_name in workbook.sheetnames:
            return sheet_name
    return workbook.sheetnames[0]


def _read_records(path: Path, index_id: str) -> tuple[str, list[str], list[str], list[dict[str, Any]]]:
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=True)
    try:
        sheet_name = _sheet_name_for(index_id, workbook)
        sheet = workbook[sheet_name]
        header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=False), ())
        headers: list[str] = []
        normalized_headers: list[str] = []
        for idx, cell in enumerate(header_cells):
            header = str(cell.value).strip() if getattr(cell, "value", None) is not None else ""
            if not header:
                header = f"col_{idx + 1}"
            headers.append(header)
            normalized_headers.append(header.lower())

        metadata_columns = _metadata_columns_for(index_id, headers)
        metadata_column_set = {column.lower() for column in metadata_columns}
        records: list[dict[str, Any]] = []

        for row_index, cells in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            if not cells:
                continue

            lexical_data: dict[str, str] = {}
            metadata: dict[str, str] = {}

            for idx, cell in enumerate(cells):
                key = normalized_headers[idx] if idx < len(normalized_headers) else f"col_{idx + 1}"
                value = _sanitize_cell(getattr(cell, "value", None))
                lexical_data[key] = value

                hyperlink = getattr(cell, "hyperlink", None)
                if hyperlink is not None:
                    target = str(getattr(hyperlink, "target", "") or getattr(hyperlink, "location", "") or "").strip()
                    if target and key == "link":
                        lexical_data[key] = target

                if key in metadata_column_set and value:
                    metadata[key] = value

            text = lexical_data.get("text", "").strip()
            if not text:
                continue

            records.append(
                {
                    "row": row_index,
                    "text": text,
                    "metadata": metadata,
                    "data": lexical_data,
                }
            )

        return sheet_name, headers, metadata_columns, records
    finally:
        workbook.close()


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _write_chunks(index_id: str, chunks: list[dict[str, Any]]) -> None:
    chunks_path = SEMANTIC_DIR / index_id / "chunks.json"
    chunks_path.parent.mkdir(parents=True, exist_ok=True)
    chunks_path.write_text(json.dumps(chunks, ensure_ascii=False, indent=2), encoding="utf-8")


def _build_records_from_raw_json(config: dict[str, Any]) -> dict[str, Any] | None:
    raw_path = Path(config["raw_path"])
    if not raw_path.exists():
        return None

    raw_rows = _load_json(raw_path)
    if not isinstance(raw_rows, list):
        raise ValueError(f"Fonte JSON bruta invalida: {raw_path}")

    records: list[dict[str, Any]] = []
    for position, raw in enumerate(raw_rows, start=1):
        if not isinstance(raw, dict):
            continue
        text = str(raw.get("text") or "").strip()
        if not text:
            continue

        date = str(raw.get("date") or "").strip()
        number = str(position)
        title = f"{config['index_label']} - {date}" if date else str(config["index_label"])
        row_map = {
            "text": text,
            "title": title,
            "date": date,
            "number": number,
            "cm": str(bool(raw.get("cm"))).lower(),
            "quest": str(bool(raw.get("quest"))).lower(),
            "pens": str(bool(raw.get("pens"))).lower(),
        }
        metadata = {
            "title": title,
            "date": date,
            "number": number,
            "cm": bool(raw.get("cm")),
            "quest": bool(raw.get("quest")),
            "pens": bool(raw.get("pens")),
        }
        records.append(
            {
                "row": position + 1,
                "text": text,
                "metadata": metadata,
                "data": row_map,
            }
        )

    semantic_chunks = rechunk_semantic_rows(records, index_label=str(config["index_label"]))
    source_relative = raw_path.relative_to(PROJECT_ROOT).as_posix()
    index_id = str(config["index_id"])
    output_name = str(config["output_name"])
    chunks = [{k: v for k, v in row.items() if k != "embedding_text"} for row in semantic_chunks]
    payload = {
        "schema_version": 1,
        "index_id": index_id,
        "index_label": str(config["index_label"]),
        "book_code": str(config["book_code"]),
        "file_stem": str(config["file_stem"]),
        "source_file": source_relative,
        "sheet_name": "",
        "headers": ["text", "title", "date", "number", "cm", "quest", "pens"],
        "text_column": "text",
        "metadata_columns": list(config["metadata_columns"]),
        "generated_at": _utc_iso_now(),
        "source_rows": len(records),
        "semantic_rows": len(chunks),
        "records": records,
    }
    _write_json(SOURCE_DIR / output_name, payload)
    _write_chunks(index_id, chunks)
    return {
        "index_id": index_id,
        "index_label": str(config["index_label"]),
        "book_code": str(config["book_code"]),
        "file_stem": str(config["file_stem"]),
        "path": output_name,
        "source_file": source_relative,
        "source_rows": len(records),
        "semantic_rows": len(chunks),
    }


def build_source_records() -> list[dict[str, Any]]:
    SOURCE_DIR.mkdir(parents=True, exist_ok=True)
    sources: list[dict[str, Any]] = []

    for xlsx_path in sorted(LEXICAL_DIR.glob("*.xlsx"), key=lambda item: item.stem.upper()):
        if xlsx_path.name.startswith("~$"):
            continue

        file_stem = xlsx_path.stem
        index_id = FILE_STEM_TO_INDEX_ID.get(file_stem.upper(), file_stem.lower())
        book_code = FILE_STEM_TO_BOOK_CODE.get(file_stem.upper(), file_stem.upper())
        sheet_name, headers, metadata_columns, records = _read_records(xlsx_path, index_id)
        index_label = file_stem.upper()
        raw_chunks = rechunk_semantic_rows(records, index_label=index_label)
        chunks = [{k: v for k, v in row.items() if k != "embedding_text"} for row in raw_chunks]
        source_relative = xlsx_path.relative_to(PROJECT_ROOT).as_posix()
        output_name = f"{index_id}.json"
        output_path = SOURCE_DIR / output_name

        payload = {
            "schema_version": 1,
            "index_id": index_id,
            "index_label": index_label,
            "book_code": book_code,
            "file_stem": file_stem,
            "source_file": source_relative,
            "sheet_name": sheet_name,
            "headers": headers,
            "text_column": "text",
            "metadata_columns": metadata_columns,
            "generated_at": _utc_iso_now(),
            "source_rows": len(records),
            "semantic_rows": len(chunks),
            "records": records,
        }
        _write_json(output_path, payload)
        _write_chunks(index_id, chunks)
        sources.append(
            {
                "index_id": index_id,
                "index_label": index_label,
                "book_code": book_code,
                "file_stem": file_stem,
                "path": output_name,
                "source_file": source_relative,
                "source_rows": len(records),
                "semantic_rows": len(semantic_chunks),
            }
        )

    for config in EXTRA_JSON_SOURCES:
        source = _build_records_from_raw_json(config)
        if source is not None:
            sources.append(source)

    index_payload = {
        "schema_version": 1,
        "generated_at": _utc_iso_now(),
        "sources": sources,
    }
    _write_json(SOURCE_DIR / "index.json", index_payload)
    return sources


def main() -> int:
    sources = build_source_records()
    print(f"Source records generated: {len(sources)} files in {SOURCE_DIR}")
    for source in sources:
        print(
            f"{source['index_id']}: records={source['source_rows']} "
            f"semantic_chunks={source['semantic_rows']} path={source['path']}"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
